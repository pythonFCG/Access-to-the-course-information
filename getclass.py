import requests
from requests.exceptions import ReadTimeout, ConnectionError, RequestException
from random import randint
import pickle
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os
import easygui as eg

headers = {
    'Referer':'http://jwc.zafu.edu.cn/',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36'
}

RadioButtonList1 = "学生"
#print(RadioButtonList1)

def get_post_data(number,passward,Code,value):
    data = {
    '__VIEWSTATE':value,
    'txtUserName':number,
    'Textbox1':'',
    'TextBox2':passward,
    'txtSecretCode':Code,
    'RadioButtonList1': RadioButtonList1,
    'Button1':'',
    'lbLanguage':'',
    'hidPdrs':'',
    'hidsc':'',
     '__EVENTVALIDATION':''
    }
    return data

def sign_in():
    s = requests.session()#相当于一个浏览器 记录所有信息
    #eg.msgbox('正在连接教务网....',image='wait.gif')
    try:
        response = s.get("http://115.236.84.162/default2.aspx",headers=headers,timeout=5)
        #print(response.status_code)
    except ReadTimeout:
        eg.msgbox('超时了，教务网好像蹦了。。。再试试。。（Timeout）','作者:Warms QQ:769737584',image='pawd.gif')
        #print('Timeout')
        return 0
    except ConnectionError:
        eg.msgbox('请检查你的网络！','作者:Warms QQ:769737584',image='pawd.gif')
        #print('Connection error')
        return 0
    except RequestException:
        eg.msgbox('教务网蹦了。。。。(HTTPerror)','作者:Warms QQ:769737584',image='pawd.gif')
        #print('Error')
        return 0

    res = s.get('http://115.236.84.162/CheckCode.aspx',headers=headers)
    with open('验证码.gif','wb') as f:
        f.write(res.content)
    #os.system('验证码.gif')
    msg = '请输入验证码(看不清请点击"Cancel")：'
    img = '验证码.gif'
    while True:
        Code = eg.enterbox(msg,'作者:Warms QQ:769737584',image=img)
        if Code == '':
            msg = '验证码不能为空！'
            continue
        elif Code==None:
            res = s.get('http://115.236.84.162/CheckCode.aspx', headers=headers)
            with open('验证码.gif', 'wb') as f:
                f.write(res.content)
            continue
        else:
            break


    soup = BeautifulSoup(response.text,'lxml')
    value =soup.find('input')['value']
    #print(value)
    try:
        response =s.post('http://115.236.84.162/default2.aspx',data=get_post_data(number=students_number,passward=passward, Code=Code,value=value),headers=headers)#登陆首页 响应
    except ReadTimeout:
        eg.msgbox('超时了，教务网好像蹦了。。。（Timeout）''作者:Warms QQ:769737584',image='pawd.gif')
        #print('Timeout')
        return 0
    except ConnectionError:
        eg.msgbox('请检查你的网络！''作者:Warms QQ:769737584',image='pawd.gif')
        #print('Connection error')
        return 0
    except RequestException:
        eg.msgbox('教务网蹦了。。。。(HTTPerror)''作者:Warms QQ:769737584',image='pawd.gif')
        #print('Error')
        return 0
    #print(response.headers)


    if  'xs_main.aspx'in response.url :#if 'xs_main.aspx'in response.url :6671
        #print('login Success!')
        choices = ['最新学年课表','成绩单']
        while True:
            choice = eg.indexbox('请选择查询类型：','作者：Warms',image='success.gif',choices=choices)
            #choice = input('请输入查询类型:1：课表,2:成绩:')
            #if choice in choices:
            if choice == 0:
                #eg.msgbox('正在查询中..请稍后',image='wait.gif')
                class_html = get_class_infor(s,response.text).replace(r'<br>', '\n')
                code= save_class_infor(class_html)
                return code
            elif choice == 1:
                #eg.msgbox('正在查询中..请稍后',image='wait.gif')
                grades_html = get_grates_infor(s,response.text).replace('&nbsp;',' ')
                code=save_grades_infor(grades_html)
                return code
            #else:
                #print('输入错误，无此选项！')
    #elif response.headers['Content-Length'] =='6671':
        #print('验证码错误！')
        #return 3
        #sign_in()
    else :#6672
        #print('用户名或密码错误')
        return 4

def get_grates_infor(s,text):
    header = {
        'Referer': 'http://115.236.84.162/xs_main.aspx?xh=' + students_number,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36'
    }
    payload = {
        '__EVENTTARGET':'',
        '__EVENTARGUMENT':'',
        '__VIEWSTATE': '',
        'hidLanguage':'',
        'ddlXN':'',
        'ddlXQ':'',
        'ddl_kcxz':'',
        'btn_zcj':'历年成绩'
    }
    params = get_params_data(text)  # 获取get请求参数
    grades_url = 'http://115.236.84.162/xscjcx.aspx'
    response = s.get(grades_url, headers=header, params=params)  # 成绩信息 响应体
    #print(response.url)
    soup = BeautifulSoup(response.text, 'lxml')
    value = soup.input. find_next_sibling(). find_next_sibling()['value']
    #print(value)
    payload['__VIEWSTATE']=value
    #print(response.url)
    response = s.post(response.url,headers=header,data=payload)
    #print(response.text)
    return response.text

def get_class_infor(s,text):
    header = {
        'Referer': 'http://115.236.84.162/xs_main.aspx?xh=' + students_number,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36'
    }
    params = get_params_data(text)#获取get请求参数
    class_url = 'http://115.236.84.162/xskbcx.aspx'
    response = s.get(class_url, headers=header, params=params )#课程信息 响应体
    #with open('html.pkl','wb') as  f:
    #    pickle.dump(response.text,f)
    #print(response.text)
    #print(response.url)
    return response.text

def get_params_data(text):
    soup = BeautifulSoup(text, 'lxml')
    name = soup.find(id="xhxm").string[:2]
    params = {
        'xh': students_number,
        'xm': name,
        'gnmkdm': 'N121603'
    }
    return params

def save_grades_infor(html):
    label_list = []
    string=''
    wb_g = openpyxl.Workbook()
    ws_g = wb_g.active

    soup = BeautifulSoup(html, 'lxml')
    td = soup.find(id="Datagrid1").select('td')
    for Label in td[:16]:
        label_list.append(Label.get_text())
    ws_g.append(label_list)
    count =0
    #print(td[16:])
    #------填充成绩信息
    for t in td[16:]:
        count+=1
        if count %16!=0:
            if t.get_text()!='':
                string +=(t.get_text()+'|')
        else:
            string += (t.get_text() + '|')
            #print(string)
            ws_g.append(string.split('|'))
            string=''

    #-------设置列宽
    column = ['A','B', 'C', 'D', 'E', 'F', 'G', 'H','I','J','K','L','M','N','O','P']
    for each in column:
        ws_g.column_dimensions[each].width = 20

    #-------填充颜色
    for each_style in ws_g['E2':'E57']:
        for each_cell in each_style:
            if each_cell.value != None:
                fill = PatternFill(fill_type="solid", fgColor='DDDDDD')
                if each_cell.value == '必修':
                    each_cell.fill = fill
    # ---------设置边框
    thick = Side(border_style="dotted", color="000000")
    dotted = Side(border_style="dotted", color="000000")
    border = Border(top=thick, left=thick, right=thick, bottom=dotted)
    file_dir = eg.diropenbox('下载成功！请选择保存路径：','请选择一个文件夹')
    if file_dir !=None:
        os.chdir(file_dir)
        wb_g.save('成绩单.xlsx')
        return 2
    else:
        wb_g.save('成绩单.xlsx')
        return 5

def save_class_infor(html):
    wb = openpyxl.Workbook()
    ws = wb.active
    #---------填充课表数据
    Personal_information = []
    #---------个人信息
    soup = BeautifulSoup(html, 'lxml')
    #print(soup.find('tr').find_next_sibling().find(id="Label5").string)  # 学号
    Personal_information.append(soup.find('tr').find_next_sibling().find(id="Label5").string)
    #print(soup.find('tr').find_next_sibling().find(id="Label6").string)  # 姓名
    Personal_information.append(soup.find('tr').find_next_sibling().find(id="Label6").string)
    #print(soup.find('tr').find_next_sibling().find(id="Label7").string)  # 学院
    Personal_information.append(soup.find('tr').find_next_sibling().find(id="Label7").string)
    #print(soup.find('tr').find_next_sibling().find(id="Label8").string)  # 专业
    Personal_information.append(soup.find('tr').find_next_sibling().find(id="Label8").string)
    #print(soup.find('tr').find_next_sibling().find(id="Label9").string)  # 行政班
    Personal_information.append(soup.find('tr').find_next_sibling().find(id="Label9").string)
    for i in range(1, 5):#A1~A5
        ws['A%d' % i] = Personal_information[i]

    #-------学年学期信息
    option = soup.find_all(selected='selected')
    year = option[0].string + ' 学年第 ' + option[1].string + ' 学期 '
    ws['E4'] = year

    #-------课表信息
    td = soup.find(id="Table1").select('td')
    a=str(randint(30,230))
    #color ="%s%s%s"%(a,a,a)
    #print(color)
    for t in td:
        if t.get_text()!=' 'and len(t.get_text())>4:
            time = t.get_text()[t.get_text().find('周'):t.get_text().find('节')]
            if '一' in time:
                ws['B%d' % (5 + int(time[3]))] = t.get_text()
                ws.merge_cells('B%d:B%d' % (int(time[3]) + 5, int(time[-1]) + 5))
            elif '二' in time:
                ws['C%d' % (5 + int(time[3]))] = t.get_text()
                ws.merge_cells('C%d:C%d' % (int(time[3]) + 5, int(time[-1]) + 5))
            elif '三' in time:
                ws['D%d' % (5 + int(time[3]))] = t.get_text()
                ws.merge_cells('D%d:D%d' % (int(time[3]) + 5, int(time[-1]) + 5))
            elif '四' in time:
                ws['E%d' % (5 + int(time[3]))] = t.get_text()
                ws.merge_cells('E%d:E%d' % (int(time[3]) + 5, int(time[-1]) + 5))
            elif '五' in time:
                ws['F%d' % (5 + int(time[3]))] = t.get_text()
                ws.merge_cells('F%d:F%d' % (int(time[3]) + 5, int(time[-1]) + 5))
            elif '六' in time:
                ws['G%d' % (5 + int(time[3]))] = t.get_text()
                ws.merge_cells('G%d:G%d' % (int(time[3]) + 5, int(time[-1]) + 5))
            elif '七' in time:
                ws['H%d' % (5 + int(time[3]))] = t.get_text()
                ws.merge_cells('H%d:H%d' % (int(time[3]) + 5, int(time[-1]) + 5))
    ws['B5'] = '星期一'
    ws['C5'] = '星期二'
    ws['D5'] = '星期三'
    ws['E5'] = '星期四'
    ws['F5'] = '星期五'
    ws['G5'] = '星期六'
    ws['H5'] = '星期天'
    for Section in range(1, 13):
        ws['A%d' % (Section + 5)] = '第%d节' % Section

    # ----------设置行列
    for i in range(5, 18):
        if i == 5:
            ws.row_dimensions[i].height = 20
        else:
            ws.row_dimensions[i].height = 100
    column = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
    for each in column:
        ws.column_dimensions[each].width = 20

    # ---------设置边框
    thick = Side(border_style="dotted", color="000000")
    dotted = Side(border_style="dotted", color="000000")
    border = Border(top=thick, left=thick, right=thick, bottom=dotted)

    # ---------设置单元格格式
    al = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for each_style in ws['A5':'H17']:
        for each_cell in each_style:
            each_cell.alignment = al
            each_cell.border = border
        # ---------设置填充色
    color = ['FFDD86', 'DDDDDD', 'aa9968', 'aa99C7', '66CCEE', 'FFC7CC', 'FFDDBB','FF6789','FF9841','aa9837','dd8899','6688aa']
    for each_style in ws['B6':'H17']:
        for each_cell in each_style:
            if each_cell.value != None:
                fill = PatternFill(fill_type="solid", fgColor=color[randint(0, 6)])
                each_cell.fill = fill
    file_dir = eg.diropenbox('下载成功！请选择保存路径：')
    if file_dir != None:
        os.chdir(file_dir)
        wb.save('课程表.xlsx')
        return 1
    else:
        wb.save('课程表.xlsx')
        return 5



path = os.getcwd()
msg='请输入学号和密码:'
try:
    with open('np.pkl','rb') as f:
        np = pickle.load(f)
except FileNotFoundError:
    np= []
    pass

while True:
    try:
        students_number,passward=eg.multpasswordbox(msg,'登陆',['用户名：','密码：'],values=np)
        np = [students_number,passward]
        with open('np.pkl','wb') as f:
            pickle.dump(np,f)
        if not students_number.isdigit() or students_number == '':
            msg='学号只能是数字且不为空！'
            continue
        elif passward == '':
            msg='密码不能为空！'
            continue
    except TypeError:
        break
    status = sign_in()
    if status == 1:
        option = eg.ccbox('成功下载课表~',choices=('直接打开','不打开'),image=path+'\suc.gif')
        if option == 1:
            os.system('课程表.xlsx')

            break
        else:
            break
    elif status == 2:
        option=eg.ccbox('成功下载成绩单~',choices=('直接打开','不打开'),image=path+'\suc.gif')
        if option==1:
            os.system('成绩单.xlsx')
            break
        else:
            break
  #  elif status == 3:

  #      eg.msgbox('(⊙﹏⊙)眼神不好啊！你！验证码错了！',image='pawd.gif')
   #     continue
    elif status == 4:
        eg.msgbox('输入信息有误请核对！',image='pawd.gif')
        continue
    elif status == 5:
        #eg.msgbox('目录都不选..我帮你保存到安装目录下了...',image=path+'\cancel.gif')
        break
    elif status==0:
        break
#html=html.replace(r'<br>','\n')
#save_class_infor(html)
#
